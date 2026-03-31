"""
AI기반 정보보호공시 솔루션
자산/비용 자동분류 - 실제 엑셀 컬럼 구조 반영 버전
"""
import streamlit as st
import pandas as pd
import numpy as np
import io
from db_helper import (
    init_db, load_table, insert_row, update_row,
    delete_rows, replace_all, truncate_table, bulk_update_classifications,
    reset_classifications,
    verify_user, get_all_users, add_user, delete_user, reset_password,
    export_backup, import_backup, is_db_empty,
)
from matching_engine import run_matching
from ai_classifier import classify_batch
from report_generator import generate_report
from github_backup import is_configured as gh_configured, push_backup as gh_push, pull_backup as gh_pull

# ────────────────────────────────────────
# 페이지 설정
# ────────────────────────────────────────
st.set_page_config(
    page_title="AI기반 정보보호공시 솔루션",
    page_icon="🛡️",
    layout="wide",
)
init_db()


def _auto_backup_to_github():
    """주요 작업 후 GitHub에 자동 백업 (설정된 경우에만)"""
    if gh_configured():
        try:
            gh_push(export_backup())
        except:
            pass  # 실패해도 사용자 작업을 방해하지 않음

# ────────────────────────────────────────
# 로그인 처리
# ────────────────────────────────────────
# ★ 마스터 복구 키 (Admin 비밀번호 분실 시 초기화용)
# 이 값을 변경하여 자체 복구 키를 설정하세요
MASTER_RECOVERY_KEY = "SecureReset2024!"

if "logged_in" not in st.session_state:
    st.session_state["logged_in"] = False
    st.session_state["user"] = None

if not st.session_state["logged_in"]:
    # 로그인 화면
    st.markdown("""
    <style>
        .login-container { max-width:420px; margin:5rem auto; padding:2.5rem;
                           background:white; border-radius:16px;
                           box-shadow:0 4px 20px rgba(0,0,0,0.1); }
        .login-title { text-align:center; font-size:1.8rem; font-weight:700;
                       color:#1E3A5F; margin-bottom:0.3rem; }
        .login-subtitle { text-align:center; font-size:0.95rem; color:#777;
                          margin-bottom:2rem; }
    </style>
    """, unsafe_allow_html=True)

    col_l, col_c, col_r = st.columns([1, 2, 1])
    with col_c:
        st.markdown("")
        st.markdown("")
        st.markdown('<p class="login-title">🛡️ AI기반 정보보호공시 솔루션</p>', unsafe_allow_html=True)
        st.markdown('<p class="login-subtitle">자산/비용 자동분류</p>', unsafe_allow_html=True)
        st.markdown("---")

        with st.form("login_form"):
            username = st.text_input("사용자 ID", placeholder="ID를 입력하세요")
            password = st.text_input("비밀번호", type="password", placeholder="비밀번호를 입력하세요")
            submitted = st.form_submit_button("로그인", use_container_width=True, type="primary")

            if submitted:
                if not username or not password:
                    st.error("ID와 비밀번호를 입력해 주세요.")
                else:
                    user = verify_user(username, password)
                    if user:
                        st.session_state["logged_in"] = True
                        st.session_state["user"] = user
                        st.rerun()
                    else:
                        st.error("❌ ID 또는 비밀번호가 일치하지 않습니다.")

        st.caption("초기 관리자 계정: Admin / Admin")

        # ── Admin 비밀번호 긴급 초기화 ──
        with st.expander("🔐 관리자 비밀번호를 잊으셨나요?"):
            st.caption("마스터 복구 키를 입력하면 Admin 비밀번호를 초기화할 수 있습니다.")
            with st.form("recovery_form"):
                recovery_key = st.text_input("마스터 복구 키", type="password", placeholder="복구 키를 입력하세요")
                new_admin_pw = st.text_input("새 Admin 비밀번호", type="password", placeholder="변경할 비밀번호")
                recover_btn = st.form_submit_button("🔑 비밀번호 초기화", use_container_width=True)

                if recover_btn:
                    if recovery_key == MASTER_RECOVERY_KEY:
                        if new_admin_pw and len(new_admin_pw) >= 4:
                            # Admin 계정 id 조회
                            users = get_all_users()
                            admin_row = users[users["username"] == "Admin"]
                            if not admin_row.empty:
                                reset_password(admin_row["id"].values[0], new_admin_pw)
                                st.success("✅ Admin 비밀번호가 변경되었습니다. 새 비밀번호로 로그인하세요.")
                            else:
                                st.error("Admin 계정을 찾을 수 없습니다.")
                        else:
                            st.error("새 비밀번호는 4자 이상 입력해 주세요.")
                    else:
                        st.error("❌ 마스터 복구 키가 일치하지 않습니다.")

    st.stop()  # 로그인 전에는 아래 코드 실행 안 함

# ── 로그인된 사용자 정보 ──
current_user = st.session_state["user"]
is_admin = current_user["role"] == "admin"

# ── Sleep 후 깨어남 감지 & 자동 복원 ──
if "sleep_checked" not in st.session_state:
    st.session_state["sleep_checked"] = True
    st.session_state["need_restore"] = False

    if is_db_empty():
        # GitHub 백업이 설정되어 있으면 자동 복원 시도
        if gh_configured():
            with st.spinner("🔄 GitHub에서 백업 데이터를 복원하는 중..."):
                pull_result = gh_pull()
            if pull_result["success"] and pull_result["data"]:
                try:
                    result = import_backup(pull_result["data"])
                    total = sum(result.values())
                    if total > 0:
                        st.success(f"✅ GitHub 백업에서 자동 복원 완료! ({total}건)")
                        st.session_state["need_restore"] = False
                    else:
                        st.session_state["need_restore"] = True
                except:
                    st.session_state["need_restore"] = True
            else:
                st.session_state["need_restore"] = True
        else:
            st.session_state["need_restore"] = True

if st.session_state.get("need_restore", False):
    st.warning("⚠️ **앱이 초기화되었습니다** (Streamlit Cloud Sleep 후 재시작)")

    if gh_configured():
        st.info("GitHub 백업에 데이터가 없거나 복원에 실패했습니다. 수동 백업 파일로 복원하세요.")
    else:
        st.info("💡 **자동 복원을 설정하려면**: Streamlit Cloud → Settings → Secrets에 "
                "GITHUB_TOKEN과 GITHUB_REPO를 추가하세요. (아래 가이드 참조)")

    restore_file = st.file_uploader(
        "📦 수동 백업 파일 업로드 (.json)", type=["json"], key="restore_upload"
    )
    col_r1, col_r2 = st.columns([1, 1])
    with col_r1:
        if restore_file and st.button("🔄 데이터 복원", type="primary", key="btn_restore"):
            try:
                result = import_backup(restore_file.read())
                total = sum(result.values())
                detail = " / ".join([f"{k}: {v}건" for k, v in result.items() if v > 0])
                st.success(f"✅ 복원 완료! 총 {total}건 ({detail})")
                st.session_state["need_restore"] = False
                # GitHub에도 백업 저장
                if gh_configured():
                    gh_push(export_backup())
                st.rerun()
            except Exception as e:
                st.error(f"복원 실패: {e}")
    with col_r2:
        if st.button("✖️ 닫기 (새로 시작)", key="btn_dismiss_restore"):
            st.session_state["need_restore"] = False
            st.rerun()
    st.markdown("---")

# ────────────────────────────────────────
# 스타일
# ────────────────────────────────────────
st.markdown("""
<style>
    .main-header  { font-size:1.8rem; font-weight:700; color:#1E3A5F; margin-bottom:.5rem; }
    .sub-header   { font-size:1.1rem; color:#555; margin-bottom:1.5rem; }

    /* 대시보드 통계 박스 */
    .stat-box     { background:#f0f4f8; border-radius:10px; padding:1rem 1.2rem;
                    text-align:center; border-left:4px solid #1E3A5F; margin-bottom:.5rem; }
    .stat-box-target { background:#FFF8E1; border-radius:10px; padding:1rem 1.2rem;
                    text-align:center; border-left:4px solid #F9A825; margin-bottom:.5rem; }
    .stat-num     { font-size:1.6rem; font-weight:700; color:#1E3A5F; }
    .stat-num-target { font-size:1.6rem; font-weight:700; color:#F9A825; }
    .stat-label   { font-size:.85rem; color:#666; }
    .stat-row-title { font-size:1rem; font-weight:600; color:#333; margin:1rem 0 .3rem 0; }

    /* 사이드바 배경 */
    div[data-testid="stSidebar"] { background: linear-gradient(180deg, #f8f9fa 0%, #e9ecef 100%); }

    /* 로고·타이틀 중앙 (검정색 텍스트) */
    .sidebar-logo { text-align:center; padding:1rem 0 0.2rem 0; }
    .sidebar-logo img { max-width:140px; border-radius:8px; }
    .sidebar-title { text-align:center; color:#1E3A5F; font-size:1.2rem; font-weight:700;
                     margin:0 0 0.1rem 0; padding:0; }
    .sidebar-subtitle { text-align:center; color:#333; font-size:0.9rem; font-weight:500;
                        margin:0 0 0.8rem 0; padding:0; }

    /* 라디오 메뉴 스타일 */
    div[data-testid="stSidebar"] .stRadio > label { display: none; }
    div[data-testid="stSidebar"] .stRadio > div[role="radiogroup"] {
        gap: 0 !important;
    }
    div[data-testid="stSidebar"] .stRadio > div[role="radiogroup"] > label {
        display: flex !important;
        align-items: center;
        padding: 0.65rem 1rem !important;
        margin: 2px 0 !important;
        border-radius: 8px !important;
        font-size: 0.95rem !important;
        color: #555 !important;
        background: transparent !important;
        border-left: 4px solid transparent;
        transition: all 0.15s ease;
        cursor: pointer;
    }
    div[data-testid="stSidebar"] .stRadio > div[role="radiogroup"] > label:hover {
        background: rgba(30,58,95,0.08) !important;
        color: #1E3A5F !important;
        border-left: 4px solid rgba(30,58,95,0.3);
    }
    div[data-testid="stSidebar"] .stRadio > div[role="radiogroup"] > label[data-checked="true"] {
        background: #1E3A5F !important;
        color: white !important;
        font-weight: 700 !important;
        border-left: 4px solid #4FC3F7;
        box-shadow: 0 2px 6px rgba(30,58,95,0.3);
    }
    /* 라디오 동그라미 숨기기 */
    div[data-testid="stSidebar"] .stRadio > div[role="radiogroup"] > label > div:first-child {
        display: none !important;
    }

    /* 사이드바 내 텍스트 색상 (검정) */
    div[data-testid="stSidebar"] .stMarkdown p,
    div[data-testid="stSidebar"] .stMarkdown h1,
    div[data-testid="stSidebar"] .stMarkdown h2,
    div[data-testid="stSidebar"] .stMarkdown h3,
    div[data-testid="stSidebar"] .stMarkdown h4 { color: #333; }
</style>
""", unsafe_allow_html=True)

# ────────────────────────────────────────
# 엑셀 ↔ DB 컬럼 매핑
# ────────────────────────────────────────
# 자산: 엑셀컬럼 → DB컬럼
ASSET_EXCEL_TO_DB = {
    "Asset Number": "asset_number",
    "Description": "description",
    "코스트센터": "cost_center",
    "일련번호": "serial_number",
    "코스트센터.1": "cost_center2",
    "자산클래스내역": "asset_class",
    "연중 상각비": "depreciation",
    "제외": "exclude_yn",
    "정보기술": "it_yn",
    "정보보호": "sec_yn",
}
ASSET_DB_TO_EXCEL = {v: k for k, v in ASSET_EXCEL_TO_DB.items()}

# 비용: 엑셀컬럼 → DB컬럼
COST_EXCEL_TO_DB = {
    "GL Date": "gl_date",
    "손익 센터": "profit_center",
    "코스트 센터": "cost_center",
    "Account": "account",
    "Account Name": "account_name",
    "사업 영역": "business_area",
    "문서 유형": "doc_type",
    "Description": "description",
    "금액(차변)": "amount",
    "제외": "exclude_yn",
    "정보기술": "it_yn",
    "정보보호": "sec_yn",
}
COST_DB_TO_EXCEL = {v: k for k, v in COST_EXCEL_TO_DB.items()}

# 매칭 결과 추가 컬럼
MATCH_EXTRA = {
    "match_type": "매칭유형",
    "match_score": "유사도(%)",
    "matched_desc": "매칭된 Description",
}

# 분류값 컬럼
CLASS_COLS = ["exclude_yn", "it_yn", "sec_yn"]


def excel_to_db_df(df: pd.DataFrame, mapping: dict) -> pd.DataFrame:
    """엑셀 DataFrame → DB 컬럼명으로 변환"""
    df = df.copy()
    rename = {k: v for k, v in mapping.items() if k in df.columns}
    df.rename(columns=rename, inplace=True)
    # 분류값: "O"는 유지, NaN/기타는 빈문자열
    for col in CLASS_COLS:
        if col in df.columns:
            df[col] = df[col].apply(lambda x: "O" if str(x).strip().upper() == "O" else "")
    # 나머지 NaN → 빈문자열 (문자열 컬럼)
    for col in df.columns:
        if df[col].dtype == object or col in CLASS_COLS:
            df[col] = df[col].fillna("")
    return df


def db_to_excel_df(df: pd.DataFrame, mapping: dict) -> pd.DataFrame:
    """DB DataFrame → 엑셀 컬럼명으로 변환"""
    df = df.copy()
    if "id" in df.columns:
        df.drop(columns=["id"], inplace=True)
    rename = {k: v for k, v in mapping.items() if k in df.columns}
    # 매칭 결과 컬럼도 변환
    rename.update({k: v for k, v in MATCH_EXTRA.items() if k in df.columns})
    df.rename(columns=rename, inplace=True)
    return df


# ────────────────────────────────────────
# 사이드바 메뉴
# ────────────────────────────────────────
import os
import base64

with st.sidebar:
    # ── 로고 + 타이틀 (중앙, 검정색) ──
    logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logo.png")
    if os.path.exists(logo_path):
        with open(logo_path, "rb") as f:
            logo_b64 = base64.b64encode(f.read()).decode()
        st.markdown(f'<div class="sidebar-logo"><img src="data:image/png;base64,{logo_b64}"></div>',
                    unsafe_allow_html=True)
    else:
        st.markdown('<p class="sidebar-title">🛡️ AI기반 정보보호공시 솔루션</p>', unsafe_allow_html=True)

    st.markdown('<p class="sidebar-subtitle">AI기반 정보보호공시 솔루션</p>', unsafe_allow_html=True)

    # ── 사용자 정보 & 로그아웃 ──
    user_display = current_user["name"] if current_user["name"] else current_user["username"]
    role_badge = "👑 관리자" if is_admin else "👤 사용자"
    st.markdown(f'<p style="text-align:center; color:#555; font-size:0.85rem; margin:0.3rem 0;">'
                f'{role_badge} <b>{user_display}</b></p>', unsafe_allow_html=True)

    if st.button("🚪 로그아웃", use_container_width=True, key="btn_logout"):
        st.session_state["logged_in"] = False
        st.session_state["user"] = None
        st.rerun()

    st.markdown("---")

    # ── 권한별 메뉴 구성 ──
    if is_admin:
        menu_items = ["🏠 대시보드", "📋 기준정보관리", "🔍 분석", "📊 리포트 산출", "👥 사용자관리"]
    else:
        menu_items = ["🏠 대시보드", "🔍 분석", "📊 리포트 산출"]

    menu = st.radio(
        "메뉴",
        menu_items,
        index=0,
        label_visibility="collapsed",
    )

    st.markdown("---")

    # ── OpenAI API 설정 ──
    st.markdown("#### 🤖 AI 설정")
    api_key = st.text_input("OpenAI API Key", type="password", key="openai_key",
                            help="미매칭 건 AI 검토에 사용됩니다")
    ai_model = st.selectbox("AI 모델", ["gpt-4o-mini", "gpt-4o", "gpt-3.5-turbo"], key="ai_model")

    # ── 데이터 백업 (Admin) ──
    if is_admin:
        st.markdown("---")
        st.markdown("#### 💾 데이터 백업")
        backup_data = export_backup()
        st.download_button(
            label="📦 백업 다운로드",
            data=backup_data,
            file_name=f"backup_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.json",
            mime="application/json",
            use_container_width=True,
            key="btn_backup_dl",
            help="기준정보, 분석결과, 사용자 계정 전체를 백업합니다. Sleep 복원용으로 보관하세요.",
        )

    st.markdown("---")
    st.markdown("<small style='color:#999'>v4.1 · SQLite + Streamlit</small>", unsafe_allow_html=True)

# ═══════════════════════════════════════
# 🏠 대시보드
# ═══════════════════════════════════════
if menu == "🏠 대시보드":
    st.markdown('<p class="main-header">🏠 대시보드</p>', unsafe_allow_html=True)
    st.markdown('<p class="sub-header">시스템 현황을 한눈에 확인하세요.</p>', unsafe_allow_html=True)

    ma = load_table("master_asset")
    mc = load_table("master_cost")
    ta = load_table("target_asset")
    tc = load_table("target_cost")

    # ── 자산 행 ──
    st.markdown('<p class="stat-row-title">🖥️ 자산 (Asset)</p>', unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1:
        st.markdown(f'<div class="stat-box"><div class="stat-num">{len(ma):,}</div>'
                    f'<div class="stat-label">기준정보 - 자산</div></div>', unsafe_allow_html=True)
    with c2:
        st.markdown(f'<div class="stat-box-target"><div class="stat-num-target">{len(ta):,}</div>'
                    f'<div class="stat-label">분석대상 - 자산</div></div>', unsafe_allow_html=True)

    # ── 비용 행 ──
    st.markdown('<p class="stat-row-title">💰 비용 (Cost)</p>', unsafe_allow_html=True)
    c3, c4 = st.columns(2)
    with c3:
        st.markdown(f'<div class="stat-box"><div class="stat-num">{len(mc):,}</div>'
                    f'<div class="stat-label">기준정보 - 비용</div></div>', unsafe_allow_html=True)
    with c4:
        st.markdown(f'<div class="stat-box-target"><div class="stat-num-target">{len(tc):,}</div>'
                    f'<div class="stat-label">분석대상 - 비용</div></div>', unsafe_allow_html=True)

    st.markdown("---")

    # 기준정보 분류 현황
    if not ma.empty or not mc.empty:
        st.markdown("#### 📊 기준정보 분류 현황")
        col_a, col_b = st.columns(2)
        with col_a:
            if not ma.empty:
                st.markdown("**자산 기준정보**")
                it_cnt = (ma["it_yn"] == "O").sum()
                sec_cnt = (ma["sec_yn"] == "O").sum()
                exc_cnt = (ma["exclude_yn"] == "O").sum()
                st.write(f"- 정보기술: {it_cnt}건 / 정보보호: {sec_cnt}건 / 제외: {exc_cnt}건")
        with col_b:
            if not mc.empty:
                st.markdown("**비용 기준정보**")
                it_cnt = (mc["it_yn"] == "O").sum()
                sec_cnt = (mc["sec_yn"] == "O").sum()
                exc_cnt = (mc["exclude_yn"] == "O").sum()
                st.write(f"- 정보기술: {it_cnt}건 / 정보보호: {sec_cnt}건 / 제외: {exc_cnt}건")

    st.markdown("---")
    st.info("💡 **사용 방법**: 좌측 메뉴에서 **기준정보관리**로 마스터 데이터를 등록한 뒤, "
            "**분석** 메뉴에서 신규 데이터를 업로드하여 자동 분류를 실행하세요.")

    st.markdown("""
#### 시스템 구조
1. **기준정보관리** — 자산·비용 마스터 데이터 CRUD 및 엑셀 업로드  
2. **분석** — 신규 데이터를 마스터 Description과 비교하여 자동 분류  
   - 1차: 전처리 후 **완전일치** 매칭  
   - 2차: 기준정보 Description이 분석대상에 포함되면 **포함매칭**  
   - 3차: RapidFuzz **유사도 매칭** (임계값 이상)  
   - 4차: 미매칭 건에 대해 **OpenAI AI 자동 분류**  
     · [KISA 정보보호 공시 가이드라인(2024)](https://isds.kisa.or.kr/kr/bbs/view.do?bbsId=B0000011&menuNo=204948&nttId=38) 자산분류표 기준 적용  
     · 정보기술/정보보호/제외 판단 + 신뢰도·판단근거 제공  
3. **결과 다운로드** — 분류 완료 데이터를 엑셀로 다운로드 (매칭유형별 색상 구분)  
    """)


# ═══════════════════════════════════════
# 📋 기준정보관리
# ═══════════════════════════════════════
elif menu == "📋 기준정보관리":
    st.markdown('<p class="main-header">📋 기준정보관리</p>', unsafe_allow_html=True)
    st.markdown('<p class="sub-header">자산 및 비용 마스터 데이터를 관리합니다.</p>', unsafe_allow_html=True)

    tab_asset, tab_cost = st.tabs(["🖥️ 자산 (Asset)", "💰 비용 (Cost)"])

    def render_master_tab(tab_key, table_name, excel_to_db, db_to_excel):
        """마스터 탭 공통 렌더링"""
        db_cols = list(excel_to_db.values())

        # ── 엑셀 업로드 ──
        st.markdown("##### 📤 엑셀 업로드 (전체 대체)")
        uploaded = st.file_uploader(
            f"{tab_key} 엑셀 파일 (.xlsx)", type=["xlsx"], key=f"upload_{tab_key}"
        )
        if uploaded:
            try:
                df_raw = pd.read_excel(uploaded)
                st.markdown(f"**읽어온 컬럼:** `{list(df_raw.columns)}`")
                df_db = excel_to_db_df(df_raw, excel_to_db)
                # DB에 있는 컬럼만 유지
                valid = [c for c in db_cols if c in df_db.columns]
                if not valid or "description" not in valid:
                    st.error("❌ 필수 컬럼(Description)이 없습니다. 엑셀 컬럼명을 확인해 주세요.")
                else:
                    df_db = df_db[valid]
                    # 미리보기 (엑셀 컬럼명으로 표시)
                    preview = db_to_excel_df(df_db, db_to_excel)
                    st.dataframe(preview.head(10), use_container_width=True)
                    st.caption(f"총 {len(df_db)}건")

                    if st.button(f"✅ {tab_key} 데이터 전체 대체", key=f"replace_{tab_key}"):
                        replace_all(table_name, df_db)
                        _auto_backup_to_github()
                        st.success(f"✅ {len(df_db)}건이 업로드되었습니다.")
                        st.rerun()
            except Exception as e:
                st.error(f"파일 처리 오류: {e}")

        st.markdown("---")

        # ── 데이터 조회 ──
        st.markdown("##### 📊 데이터 조회")
        df = load_table(table_name)
        if df.empty:
            st.warning("등록된 데이터가 없습니다. 엑셀 파일을 업로드해 주세요.")
        else:
            display = db_to_excel_df(df, db_to_excel)
            st.dataframe(display, use_container_width=True, height=400)
            st.caption(f"총 {len(df)}건")

            # ── 엑셀 다운로드 ──
            dl_buf = io.BytesIO()
            with pd.ExcelWriter(dl_buf, engine="openpyxl") as writer:
                display.to_excel(writer, index=False, sheet_name="기준정보")
                ws = writer.sheets["기준정보"]
                from openpyxl.styles import PatternFill, Font, Alignment
                hdr_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
                hdr_font = Font(color="FFFFFF", bold=True, size=10)
                for cell in ws[1]:
                    cell.fill = hdr_fill
                    cell.font = hdr_font
                    cell.alignment = Alignment(horizontal="center")
                for col_cells in ws.columns:
                    max_len = max(len(str(c.value or "")) for c in col_cells)
                    ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 40)
            dl_buf.seek(0)
            dl_filename = f"{tab_key}_기준정보.xlsx"
            st.download_button(
                label=f"📥 {tab_key}_기준정보 엑셀 다운로드",
                data=dl_buf,
                file_name=dl_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"dl_master_{tab_key}",
            )

        st.markdown("---")

        # ── 수동 추가 ──
        st.markdown("##### ➕ 데이터 추가")
        with st.expander("새 행 추가", expanded=False):
            new_data = {}
            input_cols = st.columns(3)
            for i, db_col in enumerate(db_cols):
                excel_name = db_to_excel.get(db_col, db_col)
                with input_cols[i % 3]:
                    if db_col == "depreciation" or db_col == "amount":
                        new_data[db_col] = st.number_input(excel_name, value=0.0, key=f"add_{tab_key}_{db_col}")
                    elif db_col in CLASS_COLS:
                        new_data[db_col] = st.selectbox(excel_name, ["", "O"], key=f"add_{tab_key}_{db_col}")
                    else:
                        new_data[db_col] = st.text_input(excel_name, key=f"add_{tab_key}_{db_col}")
            if st.button("➕ 추가", key=f"btn_add_{tab_key}"):
                insert_row(table_name, new_data)
                st.success("추가 완료!")
                st.rerun()

        # ── 수정 ──
        if not df.empty:
            st.markdown("##### ✏️ 데이터 수정")
            with st.expander("행 수정", expanded=False):
                row_id = st.selectbox("수정할 행 ID", df["id"].tolist(), key=f"edit_id_{tab_key}")
                if row_id:
                    row = df[df["id"] == row_id].iloc[0]
                    edit_data = {}
                    edit_cols = st.columns(3)
                    for i, db_col in enumerate(db_cols):
                        excel_name = db_to_excel.get(db_col, db_col)
                        with edit_cols[i % 3]:
                            if db_col == "depreciation" or db_col == "amount":
                                val = float(row.get(db_col, 0)) if pd.notna(row.get(db_col, 0)) else 0.0
                                edit_data[db_col] = st.number_input(excel_name, value=val, key=f"edit_{tab_key}_{db_col}")
                            elif db_col in CLASS_COLS:
                                opts = ["", "O"]
                                curr = str(row.get(db_col, ""))
                                idx = opts.index(curr) if curr in opts else 0
                                edit_data[db_col] = st.selectbox(excel_name, opts, index=idx, key=f"edit_{tab_key}_{db_col}")
                            else:
                                edit_data[db_col] = st.text_input(
                                    excel_name, value=str(row.get(db_col, "")),
                                    key=f"edit_{tab_key}_{db_col}"
                                )
                    if st.button("💾 수정 저장", key=f"btn_edit_{tab_key}"):
                        update_row(table_name, row_id, edit_data)
                        st.success("수정 완료!")
                        st.rerun()

            # ── 삭제 ──
            st.markdown("##### 🗑️ 데이터 삭제")
            with st.expander("행 삭제", expanded=False):
                del_ids = st.multiselect("삭제할 행 ID", df["id"].tolist(), key=f"del_ids_{tab_key}")
                if del_ids and st.button("🗑️ 삭제", key=f"btn_del_{tab_key}", type="primary"):
                    delete_rows(table_name, del_ids)
                    st.success(f"{len(del_ids)}건 삭제 완료!")
                    st.rerun()

    with tab_asset:
        render_master_tab("자산", "master_asset", ASSET_EXCEL_TO_DB, ASSET_DB_TO_EXCEL)

    with tab_cost:
        render_master_tab("비용", "master_cost", COST_EXCEL_TO_DB, COST_DB_TO_EXCEL)


# ═══════════════════════════════════════
# 🔍 분석
# ═══════════════════════════════════════
elif menu == "🔍 분석":
    st.markdown('<p class="main-header">🔍 분석 (자동 분류)</p>', unsafe_allow_html=True)
    st.markdown('<p class="sub-header">신규 데이터를 마스터와 비교하여 제외/정보기술/정보보호를 자동 매핑합니다.</p>',
                unsafe_allow_html=True)

    tab_a_asset, tab_a_cost = st.tabs(["🖥️ 자산 (Asset)", "💰 비용 (Cost)"])

    def render_analysis_tab(tab_key, target_table, master_table, excel_to_db, db_to_excel):
        """분석 탭 공통 렌더링"""
        db_cols = list(excel_to_db.values())

        # ────────────────────────────
        # Step 1: Upload
        # ────────────────────────────
        st.markdown("### Step 1. 분석 대상 데이터 업로드")
        st.caption("분석 대상 엑셀의 제외/정보기술/정보보호 컬럼은 비어있어도 됩니다. 분석 후 자동으로 채워집니다.")

        uploaded = st.file_uploader("분석 대상 엑셀 파일 (.xlsx)", type=["xlsx"], key=f"analysis_upload_{tab_key}")

        if uploaded:
            try:
                df_raw = pd.read_excel(uploaded)
                df_db = excel_to_db_df(df_raw, excel_to_db)
                valid = [c for c in db_cols if c in df_db.columns]

                if not valid or "description" not in valid:
                    st.error("❌ 필수 컬럼(Description)이 없습니다.")
                else:
                    for c in CLASS_COLS:
                        if c not in df_db.columns:
                            df_db[c] = ""
                    df_db = df_db[[c for c in db_cols if c in df_db.columns]]

                    preview = db_to_excel_df(df_db, db_to_excel)
                    st.dataframe(preview.head(10), use_container_width=True)
                    st.caption(f"총 {len(df_db)}건")

                    if st.button("📥 임시 테이블에 저장", type="primary", key=f"btn_save_{tab_key}"):
                        # ★ 분류값 초기화 (이전 분석 결과 또는 다운로드 파일의 값 제거)
                        df_db["exclude_yn"] = ""
                        df_db["it_yn"] = ""
                        df_db["sec_yn"] = ""
                        # ★ 매칭 컬럼을 명시적으로 추가 (DEFAULT 의존 제거)
                        df_db["match_type"] = ""
                        df_db["match_score"] = 0.0
                        df_db["matched_desc"] = ""
                        # 기존 데이터 완전 삭제 후 삽입
                        truncate_table(target_table)
                        replace_all(target_table, df_db)
                        st.success(f"✅ {len(df_db)}건이 업로드되었습니다. (이전 분석 결과 초기화)")
                        st.rerun()
            except Exception as e:
                st.error(f"파일 처리 오류: {e}")

        # 현재 임시 테이블 상태 + 초기화 버튼
        target_df = load_table(target_table)
        if not target_df.empty:
            col_info, col_reset = st.columns([3, 1])
            with col_info:
                # 이전 분석 결과 여부 확인
                has_prev = target_df["match_type"].apply(
                    lambda x: x != "" if isinstance(x, str) else False
                ).any() if "match_type" in target_df.columns else False
                status = " (분석 완료)" if has_prev else ""
                st.info(f"📌 현재 임시 테이블에 **{len(target_df)}건**의 데이터가 있습니다.{status}")
            with col_reset:
                if st.button("🗑️ 초기화", key=f"btn_reset_{tab_key}", help="임시 테이블의 모든 데이터를 삭제합니다"):
                    truncate_table(target_table)
                    st.success("임시 데이터가 초기화되었습니다.")
                    st.rerun()

        st.markdown("---")

        # ────────────────────────────
        # Step 2: Analysis
        # ────────────────────────────
        st.markdown("### Step 2. 분석 실행")

        threshold = st.slider("유사도 임계값 (Threshold)", 50, 100, 85, 5, key=f"threshold_{tab_key}")

        if st.button("🚀 분석 실행", type="primary", disabled=target_df.empty, key=f"btn_run_{tab_key}"):
            master_df = load_table(master_table)

            if master_df.empty:
                st.error("⚠️ 기준정보가 없습니다. 먼저 기준정보관리에서 마스터 데이터를 등록해 주세요.")
            else:
                with st.spinner("분석 중... Description 비교 및 유사도 매칭을 수행합니다."):
                    # ★ 분석 실행 전 DB의 이전 분류값·매칭 결과 초기화
                    reset_classifications(target_table)
                    target_df = load_table(target_table)  # 초기화된 데이터 다시 로드
                    result_df, stats = run_matching(target_df, master_df, threshold=threshold)

                # 통계
                st.markdown("#### 📊 매칭 결과 통계")
                s1, s2, s3, s4, s5 = st.columns(5)
                with s1:
                    st.metric("전체 건수", f"{stats['total']:,}")
                with s2:
                    st.metric("✅ 완전일치", f"{stats['exact']:,}")
                with s3:
                    st.metric("🟢 포함매칭", f"{stats['contain']:,}")
                with s4:
                    st.metric("🔶 유사매칭", f"{stats['fuzzy']:,}")
                with s5:
                    st.metric("❌ 미매칭", f"{stats['unmatched']:,}")

                match_rate = ((stats['exact'] + stats['contain'] + stats['fuzzy']) / stats['total'] * 100) if stats['total'] > 0 else 0
                st.progress(match_rate / 100, text=f"매칭률: {match_rate:.1f}%")

                # DB 업데이트
                updates = []
                for _, row in result_df.iterrows():
                    updates.append({
                        "id": row["id"],
                        "it_yn": row.get("it_yn", ""),
                        "sec_yn": row.get("sec_yn", ""),
                        "exclude_yn": row.get("exclude_yn", ""),
                        "match_type": row.get("match_type", ""),
                        "match_score": row.get("match_score", 0),
                        "matched_desc": row.get("matched_desc", ""),
                    })
                bulk_update_classifications(target_table, updates)
                _auto_backup_to_github()
                st.success("✅ 분석 완료! 결과가 저장되었습니다.")
                st.rerun()

        # 분석 결과 표시
        result_df = load_table(target_table)
        if not result_df.empty and "match_type" in result_df.columns:
            has_results = result_df["match_type"].apply(lambda x: x != "" if isinstance(x, str) else False).any()
            if has_results:
                st.markdown("#### 📋 분석 결과 미리보기")

                filter_type = st.multiselect(
                    "매칭 유형 필터",
                    ["완전일치", "포함매칭", "유사매칭", "AI분류", "미매칭"],
                    default=["완전일치", "포함매칭", "유사매칭", "AI분류", "미매칭"],
                    key=f"filter_{tab_key}",
                )
                view_df = result_df[result_df["match_type"].isin(filter_type)].copy()

                # 엑셀 컬럼명으로 변환 + 매칭 컬럼 추가
                display = db_to_excel_df(view_df, db_to_excel)

                # 분류·매칭 결과 컬럼을 오른쪽 끝으로 재배치
                class_col_names = ["제외", "정보기술", "정보보호"]
                match_col_names = ["매칭유형", "유사도(%)", "매칭된 Description"]
                result_cols = class_col_names + match_col_names
                other_cols = [c for c in display.columns if c not in result_cols]
                ordered_cols = other_cols + [c for c in result_cols if c in display.columns]
                display = display[ordered_cols]

                # 컬럼 그룹별 + 매칭유형별 복합 스타일링
                CLASS_BG = "#E8F5E9"
                MATCH_BG = "#EDE7F6"
                ROW_COLORS = {
                    "포함매칭": {"class": "#E0F2F1", "match": "#E0F2F1"},   # 연한 청록
                    "유사매칭": {"class": "#FFF3CD", "match": "#FFF3CD"},   # 노란색
                    "미매칭":  {"class": "#F8D7DA", "match": "#F8D7DA"},   # 빨간색
                    "AI분류":  {"class": "#D6E9FF", "match": "#D6E9FF"},   # 파란색
                }

                def highlight_cells(row):
                    mt = row.get("매칭유형", "")
                    styles = []
                    for col in row.index:
                        if col in class_col_names:
                            if mt in ROW_COLORS:
                                styles.append(f"background-color: {ROW_COLORS[mt]['class']}; font-weight: bold")
                            else:
                                styles.append(f"background-color: {CLASS_BG}; font-weight: bold")
                        elif col in match_col_names:
                            if mt in ROW_COLORS:
                                styles.append(f"background-color: {ROW_COLORS[mt]['match']}")
                            else:
                                styles.append(f"background-color: {MATCH_BG}")
                        else:
                            styles.append("")
                    return styles

                styled = display.style.apply(highlight_cells, axis=1)
                st.dataframe(styled, use_container_width=True, height=500)
                st.caption(f"표시 건수: {len(view_df)}건  |  "
                          f"⬜ 원본데이터  🟩 분류결과(제외/정보기술/정보보호)  🟪 매칭정보")
                st.caption(f"행 색상 — 🟢 완전일치  🩵 포함매칭  🟡 유사매칭  🔴 미매칭  🔵 AI분류")

                st.markdown("---")

                # ────────────────────────────
                # Step 2-1: AI Review (미매칭 건)
                # ────────────────────────────
                unmatched_df = result_df[result_df["match_type"] == "미매칭"].copy()

                if not unmatched_df.empty:
                    st.markdown("### Step 2-1. 🤖 AI 검토 (미매칭 건)")
                    st.caption(f"미매칭 {len(unmatched_df)}건에 대해 KISA 정보보호 공시 가이드라인 기준으로 OpenAI AI 분류를 시도합니다.")

                    with st.expander(f"미매칭 {len(unmatched_df)}건 미리보기", expanded=False):
                        unmatched_display = unmatched_df[["id", "description"]].copy()
                        unmatched_display.columns = ["ID", "Description"]
                        st.dataframe(unmatched_display, use_container_width=True, hide_index=True)

                    if not api_key:
                        st.warning("⚠️ 좌측 사이드바에서 **OpenAI API Key**를 입력해야 AI 검토를 실행할 수 있습니다.")
                    else:
                        if st.button("🤖 AI 검토 실행", type="primary", key=f"btn_ai_{tab_key}"):
                            unique_descs = unmatched_df[["description"]].drop_duplicates()
                            desc_list = [{"id": i, "description": row["description"]}
                                         for i, row in unique_descs.iterrows()]

                            st.info(f"🔄 고유 Description {len(desc_list)}건에 대해 AI 분류 요청 중...")
                            progress_bar = st.progress(0, text="AI 분석 중...")

                            def update_progress(current, total):
                                progress_bar.progress(
                                    current / total,
                                    text=f"AI 분석 중... ({current}/{total})"
                                )

                            ai_results = classify_batch(
                                desc_list, api_key, model=ai_model,
                                progress_callback=update_progress
                            )
                            progress_bar.empty()

                            ai_map = {}
                            errors = []
                            for r in ai_results:
                                if r.get("error"):
                                    errors.append(f"[{r['description'][:30]}...] {r['error']}")
                                else:
                                    ai_map[r["description"]] = {
                                        "it_yn": r["it_yn"],
                                        "sec_yn": r["sec_yn"],
                                        "exclude_yn": r["exclude_yn"],
                                        "confidence": r["confidence"],
                                        "reason": r["reason"],
                                    }

                            if errors:
                                with st.expander(f"⚠️ 오류 {len(errors)}건", expanded=False):
                                    for e in errors:
                                        st.text(e)

                            ai_updates = []
                            for _, row in unmatched_df.iterrows():
                                desc = row["description"]
                                if desc in ai_map:
                                    m = ai_map[desc]
                                    ai_updates.append({
                                        "id": row["id"],
                                        "it_yn": m["it_yn"],
                                        "sec_yn": m["sec_yn"],
                                        "exclude_yn": m["exclude_yn"],
                                        "match_type": "AI분류",
                                        "match_score": round(m["confidence"] * 100, 1),
                                        "matched_desc": m["reason"],
                                    })

                            if ai_updates:
                                bulk_update_classifications(target_table, ai_updates)
                                _auto_backup_to_github()
                                st.success(f"✅ AI 검토 완료! {len(ai_updates)}건이 분류되었습니다.")

                                ai_result_data = []
                                for u in ai_updates:
                                    ai_result_data.append({
                                        "Description": unmatched_df[unmatched_df["id"] == u["id"]]["description"].values[0],
                                        "제외": u["exclude_yn"],
                                        "정보기술": u["it_yn"],
                                        "정보보호": u["sec_yn"],
                                        "신뢰도": f"{u['match_score']}%",
                                        "AI 판단근거": u["matched_desc"],
                                    })
                                st.markdown("#### 🤖 AI 분류 결과")
                                ai_display = pd.DataFrame(ai_result_data)

                                def highlight_ai(row):
                                    return ["background-color: #D6E9FF"] * len(row)

                                st.dataframe(
                                    ai_display.style.apply(highlight_ai, axis=1),
                                    use_container_width=True, hide_index=True,
                                )
                                st.rerun()
                            else:
                                st.warning("AI 분류 결과가 없습니다.")

                st.markdown("---")

                # ────────────────────────────
                # Step 3: Download
                # ────────────────────────────
                st.markdown("### Step 3. 결과 다운로드")
                st.caption("분류 결과가 반영된 엑셀 파일을 다운로드합니다.")

                dl_df = db_to_excel_df(result_df, db_to_excel)
                dl_result_cols = ["제외", "정보기술", "정보보호", "매칭유형", "유사도(%)", "매칭된 Description"]
                dl_other_cols = [c for c in dl_df.columns if c not in dl_result_cols]
                dl_ordered = dl_other_cols + [c for c in dl_result_cols if c in dl_df.columns]
                dl_df = dl_df[dl_ordered]

                buffer = io.BytesIO()
                with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                    dl_df.to_excel(writer, index=False, sheet_name="분석결과")

                    ws = writer.sheets["분석결과"]
                    from openpyxl.styles import PatternFill, Font, Alignment

                    header_data_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
                    header_class_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
                    header_match_fill = PatternFill(start_color="5E35B1", end_color="5E35B1", fill_type="solid")
                    header_font = Font(color="FFFFFF", bold=True, size=10)

                    dl_class_cols = {"제외", "정보기술", "정보보호"}
                    dl_match_cols = {"매칭유형", "유사도(%)", "매칭된 Description"}

                    col_group = {}
                    for cell in ws[1]:
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center")
                        if cell.value in dl_class_cols:
                            cell.fill = header_class_fill
                            col_group[cell.column] = "class"
                        elif cell.value in dl_match_cols:
                            cell.fill = header_match_fill
                            col_group[cell.column] = "match"
                        else:
                            cell.fill = header_data_fill
                            col_group[cell.column] = "data"

                    cell_class_bg = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
                    cell_match_bg = PatternFill(start_color="EDE7F6", end_color="EDE7F6", fill_type="solid")
                    contain_teal = PatternFill(start_color="E0F2F1", end_color="E0F2F1", fill_type="solid")
                    yellow = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                    red = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                    ai_blue = PatternFill(start_color="D6E9FF", end_color="D6E9FF", fill_type="solid")

                    match_col_idx = None
                    for i, cell in enumerate(ws[1], 1):
                        if cell.value == "매칭유형":
                            match_col_idx = i
                            break

                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                        mt_val = row[match_col_idx - 1].value if match_col_idx else ""
                        if mt_val == "포함매칭":
                            row_fill = contain_teal
                        elif mt_val == "유사매칭":
                            row_fill = yellow
                        elif mt_val == "미매칭":
                            row_fill = red
                        elif mt_val == "AI분류":
                            row_fill = ai_blue
                        else:
                            row_fill = None

                        for cell in row:
                            group = col_group.get(cell.column, "data")
                            if group == "class":
                                cell.fill = row_fill if row_fill else cell_class_bg
                                cell.font = Font(bold=True)
                            elif group == "match":
                                cell.fill = row_fill if row_fill else cell_match_bg

                    for col_cells in ws.columns:
                        max_len = max(len(str(c.value or "")) for c in col_cells)
                        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 40)

                buffer.seek(0)
                filename = f"분석결과_{tab_key}.xlsx"

                st.download_button(
                    label="📥 엑셀 다운로드",
                    data=buffer,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    key=f"btn_dl_{tab_key}",
                )

    with tab_a_asset:
        render_analysis_tab("자산", "target_asset", "master_asset", ASSET_EXCEL_TO_DB, ASSET_DB_TO_EXCEL)

    with tab_a_cost:
        render_analysis_tab("비용", "target_cost", "master_cost", COST_EXCEL_TO_DB, COST_DB_TO_EXCEL)


# ═══════════════════════════════════════
# 📊 리포트 산출
# ═══════════════════════════════════════
elif menu == "📊 리포트 산출":
    st.markdown('<p class="main-header">📊 정보보호공시 리포트 산출</p>', unsafe_allow_html=True)
    st.markdown('<p class="sub-header">KISA 정보보호 공시 가이드라인(2024.03) 기반 분석 리포트를 생성합니다.</p>',
                unsafe_allow_html=True)

    # 분석 데이터 확인
    asset_df = load_table("target_asset")
    cost_df = load_table("target_cost")

    asset_analyzed = False
    cost_analyzed = False

    if not asset_df.empty and "match_type" in asset_df.columns:
        asset_analyzed = asset_df["match_type"].apply(lambda x: x != "" if isinstance(x, str) else False).any()
    if not cost_df.empty and "match_type" in cost_df.columns:
        cost_analyzed = cost_df["match_type"].apply(lambda x: x != "" if isinstance(x, str) else False).any()

    # 데이터 현황 표시
    st.markdown("#### 📋 분석 데이터 현황")
    col1, col2 = st.columns(2)
    with col1:
        if asset_analyzed:
            it_cnt = (asset_df["it_yn"] == "O").sum()
            sec_cnt = (asset_df["sec_yn"] == "O").sum()
            exc_cnt = (asset_df["exclude_yn"] == "O").sum()
            st.success(f"🖥️ **자산 분석 완료** ({len(asset_df)}건)")
            st.write(f"정보기술: {it_cnt}건 / 정보보호: {sec_cnt}건 / 제외: {exc_cnt}건")
        else:
            st.warning("🖥️ 자산 분석 미완료 — 분석 메뉴에서 먼저 분석을 실행하세요.")

    with col2:
        if cost_analyzed:
            it_cnt = (cost_df["it_yn"] == "O").sum()
            sec_cnt = (cost_df["sec_yn"] == "O").sum()
            exc_cnt = (cost_df["exclude_yn"] == "O").sum()
            st.success(f"💰 **비용 분석 완료** ({len(cost_df)}건)")
            st.write(f"정보기술: {it_cnt}건 / 정보보호: {sec_cnt}건 / 제외: {exc_cnt}건")
        else:
            st.warning("💰 비용 분석 미완료 — 분석 메뉴에서 먼저 분석을 실행하세요.")

    st.markdown("---")

    if not asset_analyzed and not cost_analyzed:
        st.error("⚠️ 자산 또는 비용 중 최소 하나의 분석이 완료되어야 리포트를 생성할 수 있습니다.")
    else:
        st.markdown("#### 📄 리포트 구성")
        st.markdown("""
        리포트에 포함되는 내용:
        1. **Executive Summary** — 총 IT 투자 / 정보보호 투자 요약, 핵심 지표
        2. **투자 현황 Overview** — 자산·비용별 IT/보안/제외 투자 현황
        3. **세부 분류별 Breakdown** — 자산분류별·계정별 상세 내역
        4. **주요 투자 항목 Top 10** — 금액 기준 상위 항목
        5. **분류 기준 및 판단 로직** — KISA 가이드라인 기준, 자동 분류 로직 설명 (Audit 대응)
        6. **회계 정합성 및 산출 방법** — 데이터 출처, 산출 기준
        7. **매칭 결과 통계** — 완전일치/포함/유사/AI/미매칭 통계
        8. **리스크 및 개선사항** — 보안 투자 비율, 미분류 항목 등
        """)

        st.markdown("---")

        if st.button("📊 리포트 생성 및 다운로드", type="primary", use_container_width=True):
            with st.spinner("리포트 생성 중... KISA 가이드라인 기반 분석 리포트를 작성합니다."):
                report_buffer = generate_report(
                    asset_df if asset_analyzed else pd.DataFrame(),
                    cost_df if cost_analyzed else pd.DataFrame(),
                )

            st.success("✅ 리포트가 생성되었습니다!")

            st.download_button(
                label="📥 리포트 다운로드 (Word)",
                data=report_buffer,
                file_name="정보보호공시_분석리포트.docx",
                mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                type="primary",
                use_container_width=True,
            )


# ═══════════════════════════════════════
# 👥 사용자관리 (Admin 전용)
# ═══════════════════════════════════════
elif menu == "👥 사용자관리" and is_admin:
    st.markdown('<p class="main-header">👥 사용자관리</p>', unsafe_allow_html=True)
    st.markdown('<p class="sub-header">시스템 사용자 계정을 관리합니다. (관리자 전용)</p>', unsafe_allow_html=True)

    # ── 사용자 목록 ──
    st.markdown("#### 📋 등록된 사용자")
    users_df = get_all_users()
    if not users_df.empty:
        display_users = users_df.copy()
        display_users["role"] = display_users["role"].map({"admin": "👑 관리자", "user": "👤 사용자"})
        display_users.columns = ["ID", "사용자 ID", "권한", "이름"]
        st.dataframe(display_users, use_container_width=True, hide_index=True)
        st.caption(f"총 {len(users_df)}명")
    else:
        st.warning("등록된 사용자가 없습니다.")

    st.markdown("---")

    # ── 사용자 추가 ──
    st.markdown("#### ➕ 사용자 추가")
    with st.form("add_user_form", clear_on_submit=True):
        col1, col2 = st.columns(2)
        with col1:
            new_username = st.text_input("사용자 ID *", placeholder="로그인 시 사용할 ID")
            new_password = st.text_input("비밀번호 *", type="password", placeholder="초기 비밀번호")
        with col2:
            new_name = st.text_input("이름", placeholder="표시될 이름")
            new_role = st.selectbox("권한", ["user", "admin"], format_func=lambda x: "👤 사용자" if x == "user" else "👑 관리자")

        submitted = st.form_submit_button("➕ 사용자 추가", type="primary", use_container_width=True)
        if submitted:
            if not new_username or not new_password:
                st.error("사용자 ID와 비밀번호는 필수입니다.")
            elif len(new_password) < 4:
                st.error("비밀번호는 4자 이상이어야 합니다.")
            else:
                if add_user(new_username, new_password, new_role, new_name):
                    _auto_backup_to_github()
                    st.success(f"✅ 사용자 '{new_username}'이(가) 추가되었습니다.")
                    st.rerun()
                else:
                    st.error(f"❌ '{new_username}'은(는) 이미 존재하는 ID입니다.")

    st.markdown("---")

    # ── 비밀번호 초기화 / 사용자 삭제 ──
    if not users_df.empty:
        st.markdown("#### ⚙️ 사용자 관리")

        col_pw, col_del = st.columns(2)

        with col_pw:
            st.markdown("**🔑 비밀번호 초기화**")
            reset_user_id = st.selectbox(
                "대상 사용자",
                users_df["id"].tolist(),
                format_func=lambda x: f"{users_df[users_df['id']==x]['username'].values[0]} ({users_df[users_df['id']==x]['name'].values[0]})",
                key="reset_pw_user"
            )
            new_pw = st.text_input("새 비밀번호", type="password", key="new_pw_input")
            if st.button("🔑 비밀번호 변경", key="btn_reset_pw"):
                if new_pw and len(new_pw) >= 4:
                    reset_password(reset_user_id, new_pw)
                    target_name = users_df[users_df["id"] == reset_user_id]["username"].values[0]
                    _auto_backup_to_github()
                    st.success(f"✅ '{target_name}'의 비밀번호가 변경되었습니다.")
                else:
                    st.error("비밀번호는 4자 이상 입력해 주세요.")

        with col_del:
            st.markdown("**🗑️ 사용자 삭제**")
            # Admin 계정은 삭제 목록에서 제외
            non_admin = users_df[users_df["username"] != "Admin"]
            if non_admin.empty:
                st.info("삭제 가능한 사용자가 없습니다. (Admin은 삭제 불가)")
            else:
                del_user_id = st.selectbox(
                    "삭제할 사용자",
                    non_admin["id"].tolist(),
                    format_func=lambda x: f"{non_admin[non_admin['id']==x]['username'].values[0]} ({non_admin[non_admin['id']==x]['name'].values[0]})",
                    key="del_user_select"
                )
                if st.button("🗑️ 사용자 삭제", key="btn_del_user", type="primary"):
                    target_name = non_admin[non_admin["id"] == del_user_id]["username"].values[0]
                    delete_user(del_user_id)
                    _auto_backup_to_github()
                    st.success(f"✅ '{target_name}'이(가) 삭제되었습니다.")
                    st.rerun()
